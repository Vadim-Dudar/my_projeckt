import requests
import sys
from bs4 import BeautifulSoup
from openpyxl import Workbook

URL = 'https://auto.ria.com/uk/search/'
HEADERS = {
   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36',
   'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
}


def get_html(url, params=None):
   r = requests.get(url, headers=HEADERS, params=params)
   return r


def get_pages_count(html):
   soup = BeautifulSoup(html, 'html.parser')
   counter = soup.find_all('a', class_='page-link')
   if len(counter) >= 2:
      return int(counter[-1].get_text().replace(' ', ''))
   else:
      return 1


def get_soup(html):
   soup = BeautifulSoup(html, 'html.parser')
   content = soup.find_all('div', class_='content-bar')

   cars = []
   soup = BeautifulSoup(html, 'html.parser')
   for car in content:
      details = soup.find_all('li', class_='item-char')
      cars.append({
         'name': car.find('span', class_='blue bold').get_text().strip(),
         'price': car.find('span', class_='bold green size22').get_text().strip(),
         'volume': f'{details[0].get_text()} {details[-1].get_text()}',
      })
   return(cars)


def do_xlsx(data):
   wb = Workbook()
   ws = wb.active
   for index, content in enumerate(data):
      ws[f'A{index+1}'] = content['name']
      ws[f'B{index+1}'] = content['price']
      ws[f'C{index+1}'] = content['volume']
   wb.save("sample_2.xlsx")


def parce():
   html = get_html(URL)
   if html.status_code == 200:
      cars = []
      counter = get_pages_count(html.text)
      for page in range(1, counter+1):
         # print(f'\rWait, {page}/{counter}')
         sys.stdout.write(f'\rWait {page}/{counter}')
         html = get_html(URL, params={'page': page})
         cars.extend(get_soup(html.text))
      do_xlsx(cars)
      print(cars)
      print(len(cars))
   else:
      print('Something wrong!')
   

parce()